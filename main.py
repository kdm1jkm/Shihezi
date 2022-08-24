############################################################
# 연도는 1982년부터 2022년 사이로 월은 3월 부터 8월 사이
# 1. 홈팀이 MBC일 때  승 column에서 H의 개수
# 2. 홈팀이 MBC일 때  승 column에서 A의 개수
# 3. 방문팀이 MBC일 때  승 column에서 A의 개수
# 4. 방문팀이 MBC일 때  승 column에서 H의 개수
#   [1] ‘1’과 ‘3’의 합 & ‘2’과 ‘4’의 합 (따로)
# 5. 홈 팀이 MBC일 때 scorelist 짝수 항의 합
# 6. 홈 팀이 MBC일 때 scorelist 홀수 항의 합
# 7. 방문팀이 MBC일 때 scorelist 홀수 항의 합
# 8. 방문팀이 MBC일 때 scorelist 짝수 항의 합
#   [2] ‘5’와 ’7’의 합 & ‘6’과 ‘8’의 합 (따로)
############################################################


from collections import defaultdict
from datetime import datetime
import pickle
import pandas as pd
import openpyxl

import os


# constants
FILE_NAME = "data"
MONTH_RANGE = range(3, 8 + 1)
TEAM = "MBC"
OUTPUT_NAME = lambda: f"output/{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

# TEAM = input("Enter team: ")


def request2(data: pd.DataFrame, team=TEAM, month_range=MONTH_RANGE):
    data = data.loc[data["month"].isin(month_range)]

    def filter_req(data: pd.DataFrame, column: str):
        data = data.loc[data[column] == team]
        data = data[["scorelist", "year"]]

        count = defaultdict(lambda: [0, 0])
        for index, row in data.iterrows():
            odd_sum = 0
            even_sum = 0
            for i, num in enumerate(
                row["scorelist"]
                if type(row["scorelist"]) != int
                else [row["scorelist"]]
            ):
                num = int(num)
                if i % 2 == 0:
                    even_sum += num
                else:
                    odd_sum += num

            count[int(row["year"])][0] += odd_sum
            count[int(row["year"])][1] += even_sum

        return list(map(lambda item: (item[0], *item[1]), count.items()))

    h_data = filter_req(data, "홈팀")
    a_data = filter_req(data, "방문팀")

    data_1 = defaultdict(lambda: 0)
    data_2 = defaultdict(lambda: 0)

    for h in h_data:
        data_1[h[0]] += h[2]
        data_2[h[0]] += h[1]
    for a in a_data:
        data_1[a[0]] += a[1]
        data_2[a[0]] += a[2]

    def convert(d):
        return list(map(lambda item: (item[0], item[1]), d.items()))

    return convert(data_1), convert(data_2)


def request1(data: pd.DataFrame, month_range=MONTH_RANGE) -> dict:
    data = data.loc[data["month"].isin(month_range)]

    홈 = data.loc[data["승"] == "H"]
    원정 = data.loc[data["승"] == "A"]

    홈_승리 = 홈.rename(columns={"홈팀": "팀"})
    원정_승리 = 원정.rename(columns={"방문팀": "팀"})

    홈_패배 = 원정.rename(columns={"홈팀": "팀"})
    원정_패배 = 홈.rename(columns={"방문팀": "팀"})

    def restructure_frame(data):
        return data[["팀", "year"]]

    count: dict[str, dict[int, list[int, int]]] = defaultdict(
        lambda: defaultdict(lambda: [0, 0])
    )

    for 승리 in [홈_승리, 원정_승리]:
        for _, row in 승리.iterrows():
            count[row["팀"]][row["year"]][0] += 1

    for 패배 in [홈_패배, 원정_패배]:
        for _, row in 패배.iterrows():
            count[row["팀"]][row["year"]][1] += 1

    result = dict()
    for k, v in count.items():
        result[k] = [(nk, nv[0], nv[1]) for nk, nv in v.items()]

    return result


def main():
    # 파일 읽어옴
    with open(FILE_NAME, "rb") as f:
        data = pd.DataFrame(pickle.load(f))

    # 원하는 결과 처라
    results = [None for _ in range(4)]
    results[0], results[1] = request1(data)
    results[2], results[3] = request2(data)

    # 파일로 저장
    wb = openpyxl.Workbook()

    def write_excel(ws, data):
        for d in data:
            ws.append(d)

    ws = wb.active
    ws.title = "행 개수"
    ws.append(["팀", "행 개수1"])
    write_excel(ws, results[0])

    ws = wb.create_sheet("행 개수2")
    ws.append(["팀", "행 개수"])
    write_excel(ws, results[1])

    ws = wb.create_sheet("합계1")
    ws.append(["년도", "합"])
    write_excel(ws, results[2])

    ws = wb.create_sheet("합계2")
    ws.append(["년도", "합"])
    write_excel(ws, results[3])

    if not os.path.isdir("./output"):
        os.mkdir("./output")

    wb.save(OUTPUT_NAME())

    # pandas로 나타내기
    dfs = list(map(pd.DataFrame, results))
    dfs[0].columns = ["팀", "행 개수"]
    dfs[1].columns = ["팀", "행 개수"]
    dfs[2].columns = ["년도", "합"]
    dfs[3].columns = ["년도", "합"]
    print(*dfs, sep="\n=========\n")


if __name__ == "__main__":
    main()
